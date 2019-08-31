<<<<<<< HEAD
import cv2
import numpy as np
import os
import pandas as pd
from imagePInternals import *
import pickle
from sklearn import model_selection
from sklearn.linear_model import LogisticRegression
import openpyxl

Directory = 'C:/Users/hamis/Dropbox/Coding/Hamish/Vectorizer Python/shapes'
os.chdir(Directory)

#Collects the data from the image library, and utilises the image processor
folders, labels, images, imFeatures = ['triangle', 'star', 'square', 'circle'], [], [], []
for folder in folders:
    print(folder)
    for path in os.listdir(os.getcwd() +'/'+folder):
        imageFeature = imageProcessor(folder+'/'+path)
        img = imageFeature.img
        images.append(img)
        labels.append(folder)
        imFeatures.append(imageFeature.detail())

#Sorts the data into a Train/Test ratio of 1:5 / 20%:80%
toTrain = 0
trainLabels, testLabels = [],[]
trainImages, testImages = [], []

for image, label, features in zip(images, labels, imFeatures):
    if toTrain < 5:
        #Appends the image data right next to the image itself to for the X value
        trainImages.append([image, features["Vertices"], features["Perimeter"], features["SumOfAngles"], features["DistFromCentre"]] )
        trainLabels.append(label)
        toTrain += 1
    else:
        testImages.append([image, features["Vertices"], features["Perimeter"], features["SumOfAngles"], features["DistFromCentre"]])
        testLabels.append(label)
    
        toTrain = 0

#trainImages = np.array(trainImages, dtype = object)
#testImages = np.array(testImages, dtype = object)
#trainLabels = np.array(trainLabels, dtype = object)
#testLabels = np.array(testLabels, dtype = object)

#Goes through the process of writing each data model to an excel sheet

dataModels = openpyxl.Workbook()
trainImagesSheet = dataModels.active
trainImagesSheet = dataModels.create_sheet("xd")
trainImagesSheet.title = "trainImagesSheet"

rowNumber = 1

for trainImage in trainImages:
    
    trainImagesSheet.cell(row = rowNumber, column = 1).value = (np.array2string(trainImage[0]))
    trainImagesSheet.cell(row = rowNumber, column = 2).value = str(trainImage[1])
    trainImagesSheet.cell(row = rowNumber, column = 3).value = str(trainImage[2])
    trainImagesSheet.cell(row = rowNumber, column = 4).value = str(trainImage[3])
    trainImagesSheet.cell(row = rowNumber, column = 5).value = str(trainImage[4])

    rowNumber +=1

testImagesSheet = dataModels.active
testImagesSheet = dataModels.create_sheet("xd")
testImagesSheet.title = "testImageSheet"

rowNumber = 1

for testImage in testImages:
    testImagesSheet.cell(row = rowNumber, column = 1).value = (np.array2string(trainImage[0]))
    testImagesSheet.cell(row = rowNumber, column = 2).value = str(testImage[1])
    testImagesSheet.cell(row = rowNumber, column = 3).value = str(testImage[2])
    testImagesSheet.cell(row = rowNumber, column = 4).value = str(testImage[3])
    testImagesSheet.cell(row = rowNumber, column = 5).value = str(testImage[4])

    rowNumber +=1

rowNumber = 1
trainLabelsSheet = dataModels.active
trainLabelsSheet = dataModels.create_sheet("xd")
trainLabelsSheet.title = "trainLabelsSheet"
for trainLabel in trainLabels:
    trainLabelsSheet.cell(row = rowNumber, column = 1).value = trainLabel

    rowNumber += 1
    
testLabelsSheet = dataModels.active
testLabelsSheet = dataModels.create_sheet("xd")
testLabelsSheet.title = "testLabelsSheet"
rowNumber = 1

for testLabel in testLabels:
    testLabelsSheet.cell(row = rowNumber, column = 1).value = testLabel
    rowNumber += 1

dataModels.save(filename = 'C:/Users/hamis/Documents/SVMModels/dataModels.xlsx')





=======
import cv2
import numpy as np
import os
import pandas as pd
from imagePInternals import *
import pickle
from sklearn import model_selection
from sklearn.linear_model import LogisticRegression
import openpyxl

Directory = 'C:/Users/hamis/Dropbox/Coding/Hamish/Vectorizer Python/shapes'
os.chdir(Directory)

#Collects the data from the image library, and utilises the image processor
folders, labels, images, imFeatures = ['triangle', 'star', 'square', 'circle'], [], [], []
ID = 1
for folder in folders:
    print(folder)
    for path in os.listdir(os.getcwd() +'/'+folder):
        imageFeature = imageProcessor(folder+'/'+path)
        img = imageFeature.img
        images.append("ID" + str(ID))
        labels.append(folder)
        imFeatures.append(imageFeature.detail())
        ID +=1


trainLabels = []
trainImages = []

for image, label, features in zip(images, labels, imFeatures):
    DFCSet = []
    PerimeterSet = []
    for i in range(16):
            Perimeter = features["Perimeter"]
            try:
                PerimeterSet.append(Perimeter[i])
            except:
                PerimeterSet.append(0)
        #Appends the image data right next to the image itself to for the X value
    for i in range(16):
        DFC = features["DistFromCentre"]
        try:
            DFCSet.append(DFC[i])
        except IndexError:
            DFCSet.append(0)
            continue
    InsertList = []
    InsertList.append(image)
    InsertList.append(features["Vertices"])
    InsertList.extend(PerimeterSet)
    InsertList.append(features["SumOfAngles"])
    InsertList.extend(DFCSet)
    trainImages.append(InsertList)                    
    trainLabels.append(label)
    

#trainImages = np.array(trainImages, dtype = object)
#testImages = np.array(testImages, dtype = object)
#trainLabels = np.array(trainLabels, dtype = object)
#testLabels = np.array(testLabels, dtype = object)

#Goes through the process of writing each data model to an excel sheet

dataModels = openpyxl.Workbook()
trainImagesSheet = dataModels.active
trainImagesSheet = dataModels.create_sheet("xd")
trainImagesSheet.title = "trainImagesSheet"

rowNumber = 1

for trainImage in trainImages:
    columnNum = 1
    for i in range(len(trainImage)):
        trainImagesSheet.cell(row = rowNumber, column = columnNum).value = (trainImage[i])
        columnNum += 1
    rowNumber +=1

rowNumber = 1

trainLabelsSheet = dataModels.active
trainLabelsSheet = dataModels.create_sheet("xd")
trainLabelsSheet.title = "trainLabelsSheet"
for trainLabel in trainLabels:
    trainLabelsSheet.cell(row = rowNumber, column = 1).value = trainLabel

    rowNumber += 1
    

dataModels.save(filename = 'C:/Users/hamis/Documents/SVMModels/dataModels.xlsx')





>>>>>>> 8a56d05e4b9cafd5471621254be609a74e1c712d
